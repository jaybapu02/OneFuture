"""Parsing and generation of weekly timetable Excel files.

A timetable cell contains text like ``7th — 12:30–1:15``. This module
extracts the class/grade, start time and end time, tolerating common
separators (—, –, -, |), whitespace and 12-hour clock values.
"""
import io
import re

from openpyxl import Workbook

DAY_NAMES = [
    "Monday",
    "Tuesday",
    "Wednesday",
    "Thursday",
    "Friday",
    "Saturday",
    "Sunday",
]

_EMPTY_CELLS = {"", "-", "—", "–", "–", "n/a", "na", "none", "nil", "x"}

_PERIOD_HEADER_RE = re.compile(r"^\s*period\s*(\d+)\s*$", re.IGNORECASE)
_DAY_HEADER_RE = re.compile(
    r"^\s*(day|day of week|days|week day|weekday)\s*$", re.IGNORECASE
)

_CLASS_RE = re.compile(
    r"(?:class\b\s*)?(?P<grade>\d{1,2})(?!\d)(?!\s*[:.])(?:\s*(?:st|nd|rd|th)\b)?",
    re.IGNORECASE,
)

_TIME_RE = re.compile(
    r"(?P<hour>\d{1,2})\s*[:.]\s*(?P<minute>\d{2})"
    r"\s*(?P<meridiem>a\.?m\.?|p\.?m\.?)?",
    re.IGNORECASE,
)

_SEPARATORS = re.compile(r"[\s,;|]+")


def normalize_day(text):
    """Map a day label ('Mon', 'MONDAY', 'mon ') to a full day name or None."""
    if not text:
        return None
    cleaned = re.sub(r"^mon(day)?$", "Monday", text.strip(), flags=re.IGNORECASE)
    lowered = cleaned.lower()
    for name in DAY_NAMES:
        if lowered == name.lower() or (
            len(lowered) == 3 and lowered == name[:3].lower()
        ):
            return name
    return None


def is_empty_cell(value):
    """True when a cell should be treated as 'no class assigned'."""
    if value is None:
        return True
    text = str(value).strip()
    return text.lower() in _EMPTY_CELLS or not text


def parse_time(text):
    """Parse a time value into datetime.time.

    Handles 12-hour and 24-hour values with optional am/pm. Values without a
    meridiem are assumed to be afternoon when the hour is below 7 (school
    periods are in the afternoon in the actual timetable, e.g. 12:30–1:15).
    Returns None when the text is not a time.
    """
    if not text:
        return None
    match = _TIME_RE.search(str(text))
    if not match:
        return None
    hour = int(match.group("hour"))
    minute = int(match.group("minute"))
    meridiem = (match.group("meridiem") or "").lower().strip(".")
    if hour > 23 or minute > 59:
        return None
    if meridiem:
        if meridiem.startswith("p") and hour < 12:
            hour += 12
        elif meridiem.startswith("a") and hour == 12:
            hour = 0
    elif hour < 7:
        hour += 12
    return datetime_time(hour, minute)


def datetime_time(hour, minute):
    import datetime

    return datetime.time(hour, minute)


def extract_class(text):
    """Extract (grade, display_name) from text like '7th' / 'Class 7'.

    Returns (grade, 'Class <grade>') or None when no grade is found.
    """
    if not text:
        return None
    cleaned = str(text).strip()
    match = _CLASS_RE.search(cleaned)
    if not match:
        return None
    grade = int(match.group("grade"))
    if not 1 <= grade <= 12:
        return None
    return grade, f"Class {grade}"


def parse_cell(value):
    """Parse one timetable cell.

    Accepts e.g. ``7th — 12:30–1:15``, ``5th - 2:00-2:40``,
    ``Class 8 | 3:20–4:00`` and ``12:30 pm to 1:15 pm``.

    Returns a dict with keys: raw, valid, grade, class_name, start_time,
    end_time, error. ``valid`` is False for unparseable cells; empty cells
    (—) are represented with ``valid=True`` and ``empty=True``.
    """
    if is_empty_cell(value):
        return {
            "raw": str(value or "").strip(),
            "empty": True,
            "valid": True,
            "grade": None,
            "class_name": None,
            "start_time": None,
            "end_time": None,
            "error": None,
        }

    raw = str(value).strip()

    grade_info = extract_class(raw)
    if grade_info is None:
        return {
            "raw": raw,
            "empty": False,
            "valid": False,
            "grade": None,
            "class_name": None,
            "start_time": None,
            "end_time": None,
            "error": "No class/grade found in cell.",
        }

    times = _TIME_RE.findall(raw)
    if len(times) < 2:
        return {
            "raw": raw,
            "empty": False,
            "valid": False,
            "grade": grade_info[0],
            "class_name": grade_info[1],
            "start_time": None,
            "end_time": None,
            "error": "Could not find start and end times in cell.",
        }

    start = parse_time(f"{times[0][0]}:{times[0][1]} {times[0][2]}".strip())
    end = parse_time(f"{times[1][0]}:{times[1][1]} {times[1][2]}".strip())
    if start is None or end is None:
        return {
            "raw": raw,
            "empty": False,
            "valid": False,
            "grade": grade_info[0],
            "class_name": grade_info[1],
            "start_time": None,
            "end_time": None,
            "error": "Could not parse times in cell.",
        }

    if end <= start:
        return {
            "raw": raw,
            "empty": False,
            "valid": False,
            "grade": grade_info[0],
            "class_name": grade_info[1],
            "start_time": start,
            "end_time": end,
            "error": "End time is not after start time.",
        }

    return {
        "raw": raw,
        "empty": False,
        "valid": True,
        "grade": grade_info[0],
        "class_name": grade_info[1],
        "start_time": start,
        "end_time": end,
        "error": None,
    }


def _find_header(rows, matcher):
    """Locate the first row index where at least one cell matches the header."""
    for row_idx, row in enumerate(rows[:30]):
        for col_idx, cell in enumerate(row):
            if cell is not None and matcher(str(cell).strip()):
                return row_idx, col_idx
    return None, None


def parse_workbook(workbook):
    """Parse a workbook (openpyxl Workbook) into timetable rows.

    Returns a dict with keys: periods (list of ints), rows (list of dicts:
    day, period, raw, valid, empty, grade, class_name, start_time, end_time,
    error), errors (list of str), detected (int count of valid non-empty rows).
    """
    errors = []
    periods = []
    parsed_rows = []
    sheet = None
    header_row_idx = None
    day_col = None
    period_cols = {}

    for candidate in workbook.worksheets:
        rows = [list(row) for row in candidate.iter_rows(values_only=True)]
        header_idx, day_header_col = _find_header(
            rows, lambda text: _DAY_HEADER_RE.match(text)
        )
        if header_idx is None:
            continue
        day_col = day_header_col
        candidate_period_cols = {}
        candidate_periods = []
        for col_idx, cell in enumerate(rows[header_idx]):
            if cell is None:
                continue
            match = _PERIOD_HEADER_RE.match(str(cell).strip())
            if match:
                period_num = int(match.group(1))
                candidate_period_cols[col_idx] = period_num
                candidate_periods.append(period_num)
        if not candidate_period_cols:
            continue
        sheet = candidate
        header_row_idx = header_idx
        period_cols = candidate_period_cols
        periods = sorted(candidate_periods)
        break

    if sheet is None or not period_cols:
        return {
            "periods": [],
            "rows": [],
            "errors": [
                "Could not find the timetable structure. The file must have a "
                "'Day' column and 'Period N' columns (e.g. Period 4, Period 5)."
            ],
            "detected": 0,
        }

    periods.sort()
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    for row in rows[header_row_idx + 1 :]:
        if row[day_col] is None:
            continue
        day = normalize_day(str(row[day_col]))
        if day is None:
            continue
        for col_idx, period_num in period_cols.items():
            cell_value = row[col_idx] if col_idx < len(row) else None
            if is_empty_cell(cell_value):
                continue
            parsed = parse_cell(cell_value)
            parsed["day"] = day
            parsed["period"] = period_num
            if parsed.get("empty"):
                continue
            if parsed["valid"]:
                parsed_rows.append(parsed)
            else:
                errors.append(
                    f"{day} Period {period_num}: {parsed['error']} "
                    f"('{parsed['raw']}')"
                )

    return {
        "periods": periods,
        "rows": parsed_rows,
        "errors": errors,
        "detected": len(parsed_rows),
    }


def build_template_workbook(periods=None):
    """Build a .xlsx template matching the actual timetable structure."""
    if periods is None:
        periods = [4, 5, 6, 7]

    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Timetable"

    headers = ["Day"] + [f"Period {p}" for p in periods]
    ws.append(headers)
    for day in DAY_NAMES[:5]:
        ws.append([day] + [""] * len(periods))

    sample = {4: "7th — 12:30–1:15", 6: "5th — 2:00–2:40", 7: "8th — 3:20–4:00"}
    for p in periods:
        if p in sample:
            ws.cell(row=2, column=2 + periods.index(p), value=sample[p])

    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    ws.column_dimensions["A"].width = 12
    for col in ws.columns:
        width = 22 if col[0].column > 1 else 12
        ws.column_dimensions[col[0].column_letter].width = width

    instructions = wb.create_sheet("Instructions")
    instructions.append(["Weekly Timetable Upload Guide"])
    instructions.append([])
    instructions.append(
        ["1. Edit the 'Weekly Timetable' sheet. One row per day (Monday–Friday)."]
    )
    instructions.append(
        ["2. Each period column header is used as the period number."]
    )
    instructions.append(
        ["3. Leave a cell empty or use '—' when there is no class that period."]
    )
    instructions.append(
        ["4. A filled cell should contain the class and the times, for example:"]
    )
    instructions.append(["     7th — 12:30–1:15   (class 7, 12:30 PM to 1:15 PM)"])
    instructions.append(["     5th — 2:00–2:40    (class 5, 2:00 PM to 2:40 PM)"])
    instructions.append(
        ["     Separators —, –, - and | are all accepted. Times can include am/pm."]
    )
    instructions.append(
        ["5. You may add more 'Period N' columns; they are detected automatically."]
    )
    instructions.append(["6. The timetable belongs to your school and applies to you."])
    instructions.append([])
    instructions.append(["Upload the file from 'My Timetable → Upload Weekly Timetable'."])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
