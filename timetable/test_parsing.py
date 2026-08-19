"""Tests for the Excel cell/timetable parsing (actual timetable format)."""
import datetime
import io

from django.test import TestCase

from timetable.parsing import (
    build_template_workbook,
    parse_cell,
    parse_workbook,
    parse_time,
    extract_class,
    is_empty_cell,
)


class CellParsingTests(TestCase):
    """1-4: parsing of Excel cells, class, start and end times."""

    def test_parse_full_cell_seven(self):
        result = parse_cell("7th — 12:30–1:15")
        self.assertTrue(result["valid"])
        self.assertEqual(result["class_name"], "Class 7")
        self.assertEqual(result["start_time"], datetime.time(12, 30))
        self.assertEqual(result["end_time"], datetime.time(13, 15))

    def test_parse_full_cell_five(self):
        result = parse_cell("5th — 2:00–2:40")
        self.assertTrue(result["valid"])
        self.assertEqual(result["class_name"], "Class 5")
        self.assertEqual(result["start_time"], datetime.time(14, 0))
        self.assertEqual(result["end_time"], datetime.time(14, 40))

    def test_supports_plain_hyphen_and_pipe_separators(self):
        for cell in ["5th - 2:00-2:40", "5th | 2:00 | 2:40", "5th – 2:00–2:40"]:
            result = parse_cell(cell)
            self.assertTrue(result["valid"], cell)
            self.assertEqual(result["start_time"], datetime.time(14, 0), cell)
            self.assertEqual(result["end_time"], datetime.time(14, 40), cell)

    def test_supports_class_3_rd_and_class_prefix_forms(self):
        for cell in ["3rd — 2:40–3:20", "3rd Class — 2:40–3:20", "Class 3 — 2:40–3:20"]:
            result = parse_cell(cell)
            self.assertTrue(result["valid"], cell)
            self.assertEqual(result["class_name"], "Class 3", cell)

    def test_supports_explicit_am_pm(self):
        result = parse_cell("8th 3:20 pm to 4:00 pm")
        self.assertTrue(result["valid"])
        self.assertEqual(result["start_time"], datetime.time(15, 20))
        self.assertEqual(result["end_time"], datetime.time(16, 0))

    def test_extract_class_from_variants(self):
        self.assertEqual(extract_class("7th")[0], 7)
        self.assertEqual(extract_class("Class 7")[0], 7)
        self.assertEqual(extract_class("7th Class")[0], 7)
        self.assertIsNone(extract_class("no class here"))

    def test_cell_without_class_is_invalid(self):
        result = parse_cell("12:30–1:15")
        self.assertFalse(result["valid"])
        self.assertIn("class", result["error"].lower())

    def test_cell_without_times_is_invalid(self):
        result = parse_cell("7th")
        self.assertFalse(result["valid"])

    def test_parse_time_afternoon_heuristic(self):
        self.assertEqual(parse_time("1:15"), datetime.time(13, 15))
        self.assertEqual(parse_time("12:30"), datetime.time(12, 30))
        self.assertEqual(parse_time("2:00"), datetime.time(14, 0))


class EmptyCellTests(TestCase):
    """5: empty cells (em dash, blank) are ignored."""

    def test_em_dash_is_empty(self):
        self.assertTrue(is_empty_cell("—"))
        self.assertTrue(is_empty_cell("–"))
        self.assertTrue(is_empty_cell(""))
        self.assertTrue(is_empty_cell(None))
        self.assertTrue(is_empty_cell("—"))

    def test_em_dash_cell_parses_as_empty(self):
        result = parse_cell("—")
        self.assertTrue(result["empty"])
        self.assertTrue(result["valid"])


def _actual_workbook():
    """Build a workbook with the real 12-class weekly timetable."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(build_template_workbook().read()))
    ws = wb["Weekly Timetable"]
    actual = [
        ["Monday", "7th — 12:30–1:15", None, "5th — 2:00–2:40", "8th — 3:20–4:00"],
        ["Tuesday", None, None, "3rd — 2:40–3:20", "6th — 3:20–4:00"],
        ["Wednesday", None, "4th — 2:00–2:40", "8th — 2:40–3:20", "3rd — 3:20–4:00"],
        ["Thursday", None, "7th — 2:00–2:40", None, "5th — 3:20–4:00"],
        ["Friday", None, None, "4th — 2:40–3:20", "6th — 3:20–4:00"],
    ]
    for i, row in enumerate(actual, start=2):
        for j, value in enumerate(row):
            ws.cell(row=i, column=j + 1, value=value)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


class WorkbookParsingTests(TestCase):
    """6-11: per-day counts and the total of 12 classes per week."""

    @classmethod
    def setUpTestData(cls):
        import openpyxl

        cls.result = parse_workbook(openpyxl.load_workbook(_actual_workbook()))

    def test_monday_produces_3_classes(self):
        self.assertEqual(
            sum(1 for r in self.result["rows"] if r["day"] == "Monday"), 3
        )

    def test_tuesday_produces_2_classes(self):
        self.assertEqual(
            sum(1 for r in self.result["rows"] if r["day"] == "Tuesday"), 2
        )

    def test_wednesday_produces_3_classes(self):
        self.assertEqual(
            sum(1 for r in self.result["rows"] if r["day"] == "Wednesday"), 3
        )

    def test_thursday_produces_2_classes(self):
        self.assertEqual(
            sum(1 for r in self.result["rows"] if r["day"] == "Thursday"), 2
        )

    def test_friday_produces_2_classes(self):
        self.assertEqual(
            sum(1 for r in self.result["rows"] if r["day"] == "Friday"), 2
        )

    def test_total_is_12_classes(self):
        self.assertEqual(len(self.result["rows"]), 12)
        self.assertEqual(self.result["detected"], 12)

    def test_no_errors_on_actual_data(self):
        self.assertEqual(self.result["errors"], [])

    def test_periods_detected_from_headers(self):
        self.assertEqual(self.result["periods"], [4, 5, 6, 7])

    def test_values_are_stored_individually(self):
        monday_p4 = next(
            r for r in self.result["rows"] if r["day"] == "Monday" and r["period"] == 4
        )
        self.assertEqual(monday_p4["class_name"], "Class 7")
        self.assertEqual(monday_p4["start_time"], datetime.time(12, 30))
        self.assertEqual(monday_p4["end_time"], datetime.time(13, 15))