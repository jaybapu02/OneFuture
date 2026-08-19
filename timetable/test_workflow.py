"""Tests for the real-world timetable workflow:

today's classes, data isolation, manual assignments, delete safety,
session survival across timetable replacement, and no mock data.
"""
import datetime
import io

import openpyxl

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from classes.models import SchoolClass, Subject
from sessions.models import Session, get_next_session_number
from tasks.models import Task
from timetable.importers import import_timetable_rows
from timetable.models import (
    DAYS_OF_WEEK,
    ManualClass,
    Timetable,
    TimetableOccurrenceRemoval,
)
from timetable.parsing import parse_workbook
from trainers.models import School, TrainerProfile

ACTUAL_ROWS = [
    {"day": "Monday", "period": 4, "grade": 7, "class_name": "Class 7",
     "start_time": datetime.time(12, 30), "end_time": datetime.time(13, 15)},
    {"day": "Monday", "period": 6, "grade": 5, "class_name": "Class 5",
     "start_time": datetime.time(14, 0), "end_time": datetime.time(14, 40)},
    {"day": "Monday", "period": 7, "grade": 8, "class_name": "Class 8",
     "start_time": datetime.time(15, 20), "end_time": datetime.time(16, 0)},
    {"day": "Tuesday", "period": 6, "grade": 3, "class_name": "Class 3",
     "start_time": datetime.time(14, 40), "end_time": datetime.time(15, 20)},
    {"day": "Tuesday", "period": 7, "grade": 6, "class_name": "Class 6",
     "start_time": datetime.time(15, 20), "end_time": datetime.time(16, 0)},
    {"day": "Wednesday", "period": 5, "grade": 4, "class_name": "Class 4",
     "start_time": datetime.time(14, 0), "end_time": datetime.time(14, 40)},
    {"day": "Wednesday", "period": 6, "grade": 8, "class_name": "Class 8",
     "start_time": datetime.time(14, 40), "end_time": datetime.time(15, 20)},
    {"day": "Wednesday", "period": 7, "grade": 3, "class_name": "Class 3",
     "start_time": datetime.time(15, 20), "end_time": datetime.time(16, 0)},
    {"day": "Thursday", "period": 5, "grade": 7, "class_name": "Class 7",
     "start_time": datetime.time(14, 0), "end_time": datetime.time(14, 40)},
    {"day": "Thursday", "period": 7, "grade": 5, "class_name": "Class 5",
     "start_time": datetime.time(15, 20), "end_time": datetime.time(16, 0)},
    {"day": "Friday", "period": 6, "grade": 4, "class_name": "Class 4",
     "start_time": datetime.time(14, 40), "end_time": datetime.time(15, 20)},
    {"day": "Friday", "period": 7, "grade": 6, "class_name": "Class 6",
     "start_time": datetime.time(15, 20), "end_time": datetime.time(16, 0)},
]


def _weekday(today, name):
    """The date of the current week matching the given weekday name."""
    days = [d for d, _ in DAYS_OF_WEEK]
    return today - datetime.timedelta(days=today.weekday() - days.index(name))


def make_world():
    school = School.objects.create(name="BMC Bagurai School", address="Bhadrak")
    user = User.objects.create_user(username="jaychandra", password="Pass@123")
    trainer = TrainerProfile.objects.create(
        user=user,
        school=school,
        employee_id="EMP-001",
        full_name="Jaychandra Dash",
    )
    subject = Subject.objects.create(name="Artificial Intelligence")
    for grade in range(3, 9):
        SchoolClass.objects.create(name=f"Class {grade}", section="", grade=grade)
    return school, trainer, subject


class TodayClassesTests(TestCase):
    """12: today's classes are determined from day of week + recurring timetable."""

    def setUp(self):
        self.school, self.trainer, self.subject = make_world()
        import_timetable_rows(self.trainer, ACTUAL_ROWS)
        self.today = timezone.localdate()
        self.day = self.today.strftime("%A")

    def test_dashboard_shows_only_todays_classes(self):
        self.client.login(username="jaychandra", password="Pass@123")
        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.status_code, 200)
        cards = response.context["timetable_cards"]
        expected = Timetable.objects.filter(
            trainer=self.trainer, day_of_week=self.day, is_active=True
        ).count()
        self.assertEqual(len(cards), expected)
        for card in cards:
            self.assertEqual(card["entry"].day_of_week, self.day)

    def test_dashboard_weekly_class_stat_is_dynamic(self):
        self.client.login(username="jaychandra", password="Pass@123")
        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.context["stats"]["weekly_classes"], 12)
        extra = Timetable.objects.create(
            trainer=self.trainer,
            school=self.school,
            school_class=SchoolClass.objects.get(name="Class 5"),
            subject=self.subject,
            day_of_week="Saturday",
            period=1,
            start_time="10:00",
            end_time="10:45",
            source="MANUAL",
        )
        self.client.login(username="jaychandra", password="Pass@123")
        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.context["stats"]["weekly_classes"], 13)
        extra.delete()


class IsolationTests(TestCase):
    """13: a trainer sees only their own timetable data."""

    def setUp(self):
        self.school, self.trainer, self.subject = make_world()
        user2 = User.objects.create_user(username="other", password="Pass@123")
        self.other = TrainerProfile.objects.create(
            user=user2, school=self.school, employee_id="EMP-002", full_name="Other"
        )
        import_timetable_rows(self.trainer, ACTUAL_ROWS)
        Timetable.objects.create(
            trainer=self.other,
            school=self.school,
            school_class=SchoolClass.objects.get(name="Class 6"),
            subject=self.subject,
            day_of_week="Monday",
            period=1,
            start_time="09:00",
            end_time="09:45",
            source="MANUAL",
        )

    def test_my_timetable_only_own_entries(self):
        self.client.login(username="jaychandra", password="Pass@123")
        response = self.client.get(reverse("timetable:my_timetable"))
        grid = response.context["grid"]
        own = [c for cells in grid.values() for c in cells]
        self.assertEqual(len(own), 12)
        for cell in own:
            self.assertEqual(cell["entry"].trainer, self.trainer)

    def test_cannot_access_another_trainers_entry(self):
        other_entry = Timetable.objects.get(trainer=self.other)
        self.client.login(username="jaychandra", password="Pass@123")
        response = self.client.get(reverse("timetable:class_detail", args=[other_entry.pk]))
        self.assertEqual(response.status_code, 404)

    def test_cannot_delete_another_trainers_entry(self):
        other_entry = Timetable.objects.get(trainer=self.other)
        self.client.login(username="jaychandra", password="Pass@123")
        response = self.client.post(
            reverse("timetable:recurring_remove_weekly", args=[other_entry.pk])
        )
        self.assertEqual(response.status_code, 404)
        other_entry.refresh_from_db()
        self.assertTrue(other_entry.is_active)


class ManualAssignmentTests(TestCase):
    """14-15: manual assignments can be created and deleted safely."""

    def setUp(self):
        self.school, self.trainer, self.subject = make_world()
        import_timetable_rows(self.trainer, ACTUAL_ROWS)
        self.client.login(username="jaychandra", password="Pass@123")
        self.friday = _weekday(timezone.localdate(), "Friday")

    def test_manual_assignment_created_and_labeled(self):
        response = self.client.post(
            reverse("timetable:assign_manual_class"),
            {
                "school_class": SchoolClass.objects.get(name="Class 5").pk,
                "subject": self.subject.pk,
                "date": self.friday.isoformat(),
                "period": "5",
                "start_time": "13:00",
                "end_time": "13:45",
                "notes": "Special Class 5 session",
            },
        )
        self.assertEqual(response.status_code, 302)
        manual = ManualClass.objects.get(trainer=self.trainer)
        self.assertEqual(manual.school, self.school)
        self.assertEqual(manual.school_class.name, "Class 5")
        self.assertEqual(manual.period, 5)
        self.assertEqual(Timetable.objects.filter(trainer=self.trainer, is_active=True).count(), 12)

    def test_manual_delete_only_removes_the_assignment(self):
        manual = ManualClass.objects.create(
            trainer=self.trainer,
            school=self.school,
            school_class=SchoolClass.objects.get(name="Class 5"),
            subject=self.subject,
            date=self.friday,
            period=5,
            start_time="13:00",
            end_time="13:45",
        )
        response = self.client.post(
            reverse("timetable:manual_class_delete", args=[manual.pk])
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(ManualClass.objects.filter(pk=manual.pk).exists())
        self.assertEqual(
            Timetable.objects.filter(trainer=self.trainer, is_active=True).count(),
            12,
        )

    def test_cannot_delete_another_trainers_manual_class(self):
        user2 = User.objects.create_user(username="other", password="Pass@123")
        other = TrainerProfile.objects.create(
            user=user2, school=self.school, employee_id="EMP-002", full_name="Other"
        )
        manual = ManualClass.objects.create(
            trainer=other,
            school=self.school,
            school_class=SchoolClass.objects.get(name="Class 5"),
            subject=self.subject,
            date=self.friday,
            period=5,
            start_time="13:00",
            end_time="13:45",
        )
        response = self.client.post(
            reverse("timetable:manual_class_delete", args=[manual.pk])
        )
        self.assertEqual(response.status_code, 404)
        self.assertTrue(ManualClass.objects.filter(pk=manual.pk).exists())


class DeleteSafetyTests(TestCase):
    """16: the recurring timetable is not accidentally deleted."""

    def setUp(self):
        self.school, self.trainer, self.subject = make_world()
        import_timetable_rows(self.trainer, ACTUAL_ROWS)
        self.client.login(username="jaychandra", password="Pass@123")
        self.monday = _weekday(timezone.localdate(), "Monday")
        self.entry = Timetable.objects.get(
            trainer=self.trainer, day_of_week="Monday", period=4
        )

    def test_remove_for_one_date_keeps_weekly_rule(self):
        response = self.client.post(
            reverse("timetable:occurrence_remove_date", args=[self.entry.pk]),
            {"date": self.monday.isoformat()},
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(TimetableOccurrenceRemoval.objects.filter(
            timetable=self.entry, date=self.monday
        ).exists())
        self.entry.refresh_from_db()
        self.assertTrue(self.entry.is_active)

    def test_removed_occurrence_hidden_from_dashboard(self):
        TimetableOccurrenceRemoval.objects.create(
            timetable=self.entry, date=self.monday
        )
        self.client.login(username="jaychandra", password="Pass@123")
        response = self.client.get(reverse("dashboard"))
        cards = response.context["timetable_cards"]
        if self.monday == timezone.localdate():
            self.assertNotIn(self.entry, [c["entry"] for c in cards])

    def test_remove_from_weekly_requires_confirmation(self):
        response = self.client.get(
            reverse("timetable:recurring_remove_weekly", args=[self.entry.pk])
        )
        self.assertEqual(response.status_code, 200)
        self.entry.refresh_from_db()
        self.assertTrue(self.entry.is_active)

    def test_remove_from_weekly_deactivates_only_that_entry(self):
        response = self.client.post(
            reverse("timetable:recurring_remove_weekly", args=[self.entry.pk])
        )
        self.assertEqual(response.status_code, 302)
        self.entry.refresh_from_db()
        self.assertFalse(self.entry.is_active)
        self.assertEqual(
            Timetable.objects.filter(trainer=self.trainer, is_active=True).count(),
            11,
        )


class TimetableReplacementTests(TestCase):
    """17-18: sessions survive timetable changes and numbering stays correct."""

    def setUp(self):
        self.school, self.trainer, self.subject = make_world()
        import_timetable_rows(self.trainer, ACTUAL_ROWS)
        self.monday = _weekday(timezone.localdate(), "Monday")

    def test_sessions_survive_timetable_replacement(self):
        entry = Timetable.objects.get(
            trainer=self.trainer, day_of_week="Monday", period=4
        )
        cls7 = SchoolClass.objects.get(name="Class 7")
        session = Session.objects.create(
            trainer=self.trainer,
            school=self.school,
            timetable=entry,
            school_class=cls7,
            subject=self.subject,
            session_number=1,
            date=self.monday,
            start_time=entry.start_time,
            end_time=entry.end_time,
            students_present=28,
            topic_taught="Introduction to AI",
        )
        import_timetable_rows(self.trainer, ACTUAL_ROWS, mode="replace")
        session.refresh_from_db()
        self.assertEqual(session.topic_taught, "Introduction to AI")
        self.assertIsNotNone(session.timetable)
        self.assertEqual(session.timetable.pk, entry.pk)
        self.assertEqual(session.session_number, 1)

    def test_session_numbers_do_not_reset_after_replacement(self):
        cls7 = SchoolClass.objects.get(name="Class 7")
        for i in range(1, 3):
            Session.objects.create(
                trainer=self.trainer,
                school=self.school,
                school_class=cls7,
                subject=self.subject,
                session_number=i,
                date=self.monday - datetime.timedelta(weeks=i),
                students_present=25,
                topic_taught="Topic",
            )
        import_timetable_rows(self.trainer, ACTUAL_ROWS, mode="replace")
        self.assertEqual(get_next_session_number(cls7, self.subject), 3)

    def test_tasks_survive_timetable_replacement(self):
        entry = Timetable.objects.get(
            trainer=self.trainer, day_of_week="Monday", period=4
        )
        Task.objects.create(
            trainer=self.trainer,
            school=self.school,
            timetable=entry,
            school_class=entry.school_class,
            subject=self.subject,
            title="Prepare quiz",
            date=self.monday,
            start_time=entry.start_time,
            end_time=entry.end_time,
        )
        import_timetable_rows(self.trainer, ACTUAL_ROWS, mode="replace")
        self.assertEqual(Task.objects.filter(title="Prepare quiz").count(), 1)


class NoMockDataTests(TestCase):
    """19: no mock/demo data is displayed — real empty states instead."""

    def setUp(self):
        self.school, self.trainer, self.subject = make_world()
        self.client.login(username="jaychandra", password="Pass@123")

    def test_empty_dashboard_shows_real_empty_state(self):
        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["timetable_cards"], [])
        self.assertEqual(response.context["stats"]["today_classes"], 0)
        self.assertEqual(response.context["stats"]["completed_today"], 0)
        self.assertEqual(response.context["stats"]["students_today"], 0)
        self.assertContains(response, "No classes scheduled for today")

    def test_seeded_dashboard_has_no_fabricated_names(self):
        import_timetable_rows(self.trainer, ACTUAL_ROWS)
        response = self.client.get(reverse("dashboard"))
        content = response.content.decode()
        for fake in ["Priya", "Rahul", "Jaychandra Sharma", "10 AM"]:
            self.assertNotIn(fake, content)
        self.assertContains(response, "Jaychandra Dash")


class UploadFlowTests(TestCase):
    """The upload → preview → confirm flow with the real timetable."""

    def setUp(self):
        self.school, self.trainer, self.subject = make_world()
        self.client.login(username="jaychandra", password="Pass@123")

    def _make_upload(self):
        import openpyxl

        from timetable.parsing import build_template_workbook

        wb = openpyxl.load_workbook(io.BytesIO(build_template_workbook().read()))
        ws = wb["Weekly Timetable"]
        actual = [
            ["Monday", "7th — 12:30–1:15", None, "5th — 2:00–2:40", "8th — 3:20–4:00"],
            ["Tuesday", None, None, "3rd — 2:40–3:20", "6th — 3:20–4:00"],
            ["Wednesday", None, "4th — 2:00–2:40", "8th — 2:40–3:20", "3rd — 3:20–4:00"],
            ["Thursday", None, "7th — 2:00–2:40", None, "5th — 3:20–4:00"],
            ["Friday", None, None, "4th — 2:40–3:20", "6th — 3:20–4:00"],
        ]
        for i, row in enumerate(actual, start=2):
            for j, value in enumerate(row):
                ws.cell(row=i, column=j + 1, value=value)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def test_upload_preview_detects_12_classes(self):
        buffer = self._make_upload()
        buffer.name = "timetable.xlsx"
        response = self.client.post(
            reverse("timetable:upload_timetable"), {"file": buffer}
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["detected"], 12)
        self.assertEqual(len(response.context["rows"]), 12)

    def test_confirm_import_creates_recurring_entries(self):
        buffer = self._make_upload()
        buffer.name = "timetable.xlsx"
        response = self.client.post(
            reverse("timetable:upload_timetable"), {"file": buffer}
        )
        self.assertIn("timetable_import_rows", self.client.session)
        response = self.client.post(
            reverse("timetable:upload_timetable"), {"confirm": "1"}
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            Timetable.objects.filter(trainer=self.trainer, source="EXCEL", is_active=True).count(),
            12,
        )

    def test_rejects_non_xlsx(self):
        buffer = io.BytesIO(b"not an xlsx")
        buffer.name = "timetable.xls"
        response = self.client.post(
            reverse("timetable:upload_timetable"), {"file": buffer}
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Only .xlsx files are accepted")

    def test_template_download(self):
        response = self.client.get(reverse("timetable:download_template"))
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "spreadsheetml", response["Content-Type"]
        )
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertIn("Weekly Timetable", wb.sheetnames)
        self.assertIn("Instructions", wb.sheetnames)